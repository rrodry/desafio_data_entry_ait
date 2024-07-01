import os
from datetime import datetime
from openpyxl import load_workbook, workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import requests

def file_size(direction):
    if os.path.exists(direction):
        return os.path.getsize(direction)
    return 0


import os
from datetime import datetime

current_directory = os.path.join(os.getcwd(), 'downloads')


def file_change_name(nameFiles, current):
    date_today = datetime.now().strftime("%d-%m-%Y")  # Fecha actual
    print("Archivos antes del cambio de nombre:", nameFiles)
    for nameFile in nameFiles:
        if '.' in nameFile:
            extension = nameFile.rsplit('.', 1)
            new_name = f"{extension[0]}_{date_today}.{extension[-1]}"
        else:
            new_name = f"{nameFile}_{date_today}"

        directionOld = os.path.join(current, nameFile)
        directionNew = os.path.join(current, new_name)

        os.rename(directionOld, directionNew)
        print(f"Archivo renombrado: {nameFile} -> {os.path.basename(directionNew)}")
    return 0


def modify_excel():
    date = datetime.now().strftime("%d-%m-%Y")  # Date today
    columnData = ["CODIGO", "DESCR", "MARCA", "PRECIO"]
    files = os.listdir(current_directory)
    maxCharacter = 100

    for filename in files:  # All Files
        if date in filename:  # if date is on name files
            splitFilename = filename.split(".")
            if splitFilename[1] == "xlsx":  # check if xlsx and if not this should be converted but is not ready :(
                #select file to load on work
                work = load_workbook(os.path.join(current_directory, filename))
                ws = work.active

                if "sheet1" not in work.sheetnames: #if no sheet create
                    work.create_sheet("sheet1")
                wsDestiny = work["sheet1"]
                #loop sheets on work
                for sheet in work.sheetnames:
                    if sheet != "sheet1":
                        worksheet = work[sheet]
                        columReady = []
                        #columns loop
                        for column in range(1, worksheet.max_column + 1):
                            char = get_column_letter(column)
                            firstValue = False
                            firstValueDesc = False
                            last_row_destiny = 1
                            #rows loop
                            for row in range(1, worksheet.max_row + 1):
                                valueCell = worksheet[char + str(row)].value

                                for indexColumData in range(0, len(columnData)):
                                    #if firsvalue is true conditionals for modify excel
                                    if firstValue:
                                        if valueCell is not None:
                                            if isinstance(valueCell, float):
                                                floatValue = str(valueCell).replace(",", ".")
                                                wsDestiny.cell(last_row_destiny, column, floatValue)
                                            if firstValueDesc:
                                                truncated_text = valueCell[:maxCharacter]
                                                wsDestiny.cell(last_row_destiny, column, truncated_text)
                                            if "AutoFix" in splitFilename[0]:
                                                print(sheet)
                                                wsDestiny.cell(last_row_destiny, column, sheet)
                                            wsDestiny.cell(last_row_destiny, column, valueCell)
                                            last_row_destiny += 1
                                    #if is string for know the title of the column
                                    if isinstance(valueCell, str):
                                        split = valueCell.split(" ")
                                        if split[0] in columnData:
                                            if columnData[indexColumData] in valueCell and columnData[
                                                indexColumData] not in columReady:
                                                firstValue = True
                                                if "DESCR" in valueCell:
                                                    firstValueDesc = True
                                                columReady.append(columnData[indexColumData])
                                                wsDestiny.cell(last_row_destiny, column, valueCell)
                                                wsDestiny.cell(last_row_destiny, column, "MARCA")
                                                last_row_destiny += 1
                                                break

                work.save(os.path.join(current_directory, filename))


def uploadFiles():
    url = 'https://ejemplo.com/subir-archivo'
    filesLocal = os.listdir(current_directory)
    for filename in filesLocal:
        directionFile = os.path.join(current_directory, filename)
        try:
            with open(directionFile, 'rb') as file:
                files = {'file': file}  # 'file' es el nombre del campo del formulario en el servidor
                response = requests.post(url, files=files)
                if response.status_code == 200:
                    print('Archivo subido correctamente.')
                else:
                    print(f'Error al subir el archivo. Código de estado: {response.status_code}')
        except FileNotFoundError:
            print(f'Archivo no encontrado en la ruta: {directionFile}')

